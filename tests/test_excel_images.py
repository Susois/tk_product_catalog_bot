from openpyxl import load_workbook
from PIL import Image

from app import excel
from app.models import Product


def test_append_product_downloads_all_images(tmp_path, monkeypatch) -> None:
    output = tmp_path / "products.xlsx"
    downloaded_urls: list[str] = []

    def fake_download(url, image_dir, product_id):
        downloaded_urls.append(url)
        return None

    monkeypatch.setattr(excel, "_download_image", fake_download)
    product = Product(
        source_url="https://vt.tiktok.com/example",
        resolved_url="https://www.tiktok.com/view/product/123",
        product_id="123",
        name="Product with images",
        price="10000 VND",
        image_urls=["https://img.test/1.jpg", "https://img.test/2.jpg", "https://img.test/3.jpg"],
    )

    excel.append_product(output, product)

    sheet = load_workbook(output).active
    assert downloaded_urls == product.image_urls
    assert sheet.max_row == 4
    assert sheet["B2"].value == "123"
    assert sheet["F2"].value == product.image_urls[0]
    assert sheet["F3"].value == product.image_urls[1]
    assert sheet["F4"].value == product.image_urls[2]


def test_download_converts_webp_to_real_jpeg(tmp_path, monkeypatch) -> None:
    source = tmp_path / "source.webp"
    Image.new("RGBA", (10, 10), (255, 0, 0, 128)).save(source, "WEBP")

    class Response:
        content = source.read_bytes()

        def raise_for_status(self):
            return None

    monkeypatch.setattr(excel.httpx, "get", lambda *args, **kwargs: Response())
    result = excel._download_image("https://img.test/photo.webp", tmp_path / "images", "123")

    assert result is not None
    with Image.open(result) as image:
        assert image.format == "JPEG"


def test_append_product_recovers_corrupt_workbook(tmp_path, monkeypatch) -> None:
    output = tmp_path / "products.xlsx"
    output.write_bytes(b"not an xlsx")
    monkeypatch.setattr(excel, "_download_image", lambda *args: None)
    product = Product(
        source_url="https://vt.tiktok.com/example",
        resolved_url="https://www.tiktok.com/view/product/123",
        product_id="123",
        name="Recovered product",
        price="10000",
        image_urls=[],
    )

    excel.append_product(output, product)

    assert load_workbook(output).active["B2"].value == "123"
    assert list(tmp_path.glob("products.corrupt_*.xlsx"))
