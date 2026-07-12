from __future__ import annotations

import re
import os
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image

from .models import Product

HEADERS = ["Thời gian", "Mã sản phẩm", "Tên sản phẩm", "Giá", "Tiền tệ", "Ảnh", "Link gốc", "Link TikTok Shop"]
VIETNAM_TIMEZONE = timezone(timedelta(hours=7))


class WorkbookLockedError(RuntimeError):
    """Raised when Windows prevents replacing an open Excel file."""



def _new_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sản phẩm TikTok"
    sheet.append(HEADERS)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    widths = {"A": 20, "B": 22, "C": 65, "D": 20, "E": 12, "F": 24, "G": 40, "H": 55}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _download_image(url: str, image_dir: Path, product_id: str) -> Path | None:
    if not url:
        return None
    image_dir.mkdir(parents=True, exist_ok=True)
    safe_id = re.sub(r"[^a-zA-Z0-9_-]", "_", product_id or "product")
    target = image_dir / f"{safe_id}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.jpg"
    try:
        response = httpx.get(url, follow_redirects=True, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        response.raise_for_status()
        target.write_bytes(response.content)
        # TikTok often serves WebP/AVIF bytes regardless of the URL or file
        # extension. openpyxl cannot package WebP images, so always convert the
        # downloaded image to a real RGB JPEG before embedding it.
        with Image.open(target) as image:
            image.load()
            if image.mode != "RGB":
                background = Image.new("RGB", image.size, "white")
                if "A" in image.getbands():
                    background.paste(image, mask=image.getchannel("A"))
                else:
                    background.paste(image)
                image = background
            image.save(target, format="JPEG", quality=90)
        return target
    except (httpx.HTTPError, OSError):
        target.unlink(missing_ok=True)
        return None


def _load_or_recover_workbook(path: Path):
    if not path.exists():
        _new_workbook(path)

    # Validate the Office package before openpyxl touches it. On Windows,
    # openpyxl may leave its ZipFile handle open when loading a damaged XLSX
    # raises midway, which then prevents this same process from renaming it.
    package_is_valid = False
    try:
        with zipfile.ZipFile(path) as archive:
            package_is_valid = "[Content_Types].xml" in archive.namelist()
    except (OSError, zipfile.BadZipFile):
        package_is_valid = False

    if package_is_valid:
        return load_workbook(path)

    try:
        # Preserve the broken file for inspection instead of overwriting it.
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = path.with_name(f"{path.stem}.corrupt_{timestamp}{path.suffix}")
        path.replace(backup)
    except PermissionError as error:
        raise WorkbookLockedError(
            f"File Excel '{path.name}' đang bị tiến trình bot cũ hoặc chương trình khác giữ. "
            "Hãy tắt và chạy lại bot rồi gửi lại link sản phẩm."
        ) from error
    _new_workbook(path)
    return load_workbook(path)


def _save_workbook_atomic(workbook, path: Path) -> None:
    temporary = path.with_name(f".{path.name}.tmp.xlsx")
    try:
        workbook.save(temporary)
        # Verify that the generated XLSX is a valid Office ZIP package before
        # replacing the user's current catalog.
        with zipfile.ZipFile(temporary) as archive:
            if "[Content_Types].xml" not in archive.namelist():
                raise ValueError("Generated workbook is not a valid XLSX file")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def append_product(path: Path, product: Product) -> Path:
    workbook = _load_or_recover_workbook(path)
    sheet = workbook["Sản phẩm TikTok"]
    row = sheet.max_row + 1
    sheet.append([
        datetime.now(VIETNAM_TIMEZONE).strftime("%Y-%m-%d %H:%M:%S"),
        product.product_id,
        product.name,
        product.price,
        product.currency,
        product.image_urls[0] if product.image_urls else "",
        product.source_url,
        product.resolved_url,
    ])
    image_urls = list(dict.fromkeys(url for url in product.image_urls if url))
    image_rows = max(1, len(image_urls))
    for offset in range(image_rows):
        current_row = row + offset
        if offset:
            sheet.append(["", "", "", "", "", image_urls[offset], "", ""])
        sheet.row_dimensions[current_row].height = 110
        for cell in sheet[current_row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for index, image_url in enumerate(image_urls):
        current_row = row + index
        image_path = _download_image(
            image_url,
            path.parent / "product-images",
            f"{product.product_id}_{index + 1}",
        )
        if image_path:
            image = ExcelImage(str(image_path))
            image.width = 140
            image.height = 140
            sheet.add_image(image, f"F{current_row}")
            sheet[f"F{current_row}"] = ""
    try:
        _save_workbook_atomic(workbook, path)
    except PermissionError as error:
        raise WorkbookLockedError(
            f"File Excel '{path.name}' đang được mở hoặc bị chương trình khác sử dụng. "
            "Hãy đóng file trong Microsoft Excel/Preview rồi gửi lại link sản phẩm."
        ) from error
    finally:
        workbook.close()
    return path
