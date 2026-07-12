from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook

from .models import Product

CATEGORY_FILES = {
    "quan_ao_nu": "quan_ao_nu.xlsx",
    "do_lot_nu": "do_lot_nu.xlsx",
    "phu_kien_linh_tinh": "phu_kien_linh_tinh.xlsx",
    "giay_dep": "giay_dep.xlsx",
}

CATEGORY_LABELS = {
    "quan_ao_nu": "Quần áo nữ",
    "do_lot_nu": "Đồ lót nữ",
    "phu_kien_linh_tinh": "Phụ kiện linh tinh",
    "giay_dep": "Giày dép",
}

UNDERWEAR_KEYWORDS = {
    "ao lot", "ao nguc", "bra", "bralette", "quan lot", "quan chip", "noi y",
    "lingerie", "do lot", "corset", "gen nit bung", "gen bung", "vay ngu",
}
SHOES_KEYWORDS = {
    "giay", "dep", "sandal", "sneaker", "boot", "boots", "cao got", "guoc",
    "slipper", "loafer", "mary jane", "crocs",
}
CLOTHING_KEYWORDS = {
    "ao", "quan", "vay", "dam", "chan vay", "set bo", "bo do", "jumpsuit",
    "hoodie", "cardigan", "blazer", "vest", "khoac", "jean", "legging",
    "croptop", "crop top", "polo", "thun nu", "so mi", "yem", "bikini",
}


def _plain_text(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value.lower())
    without_marks = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    return " ".join(without_marks.replace("đ", "d").split())


def _contains_keyword(text: str, keywords: set[str]) -> bool:
    return any(re.search(rf"(?<!\w){re.escape(keyword)}(?!\w)", text) for keyword in keywords)


def classify_product(name: str) -> str:
    text = _plain_text(name)
    if _contains_keyword(text, UNDERWEAR_KEYWORDS):
        return "do_lot_nu"
    if _contains_keyword(text, SHOES_KEYWORDS):
        return "giay_dep"
    if _contains_keyword(text, CLOTHING_KEYWORDS):
        return "quan_ao_nu"
    return "phu_kien_linh_tinh"


def category_excel_path(output_excel: Path, category: str) -> Path:
    return output_excel.parent / CATEGORY_FILES[category]


def normalize_url(url: str) -> str:
    value = url.strip()
    if not value:
        return ""
    parts = urlsplit(value)
    host = (parts.hostname or "").lower()
    if not host:
        return value
    path = re.sub(r"/+$", "", parts.path) or "/"
    return urlunsplit(("https", host, path, "", ""))


class ProductRegistry:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.urls: set[str] = set()
        self.product_ids: set[str] = set()
        self._load()

    def _load(self) -> None:
        if not self.path.exists():
            return
        try:
            data = json.loads(self.path.read_text(encoding="utf-8"))
            self.urls = {str(value) for value in data.get("urls", []) if value}
            self.product_ids = {str(value) for value in data.get("product_ids", []) if value}
        except (json.JSONDecodeError, OSError):
            self.urls = set()
            self.product_ids = set()

    def contains_url(self, url: str) -> bool:
        return normalize_url(url) in self.urls

    def contains_product(self, product_id: str) -> bool:
        return bool(product_id and product_id in self.product_ids)

    def register(self, product: Product) -> None:
        self.register_values(product.product_id, product.source_url, product.resolved_url)

    def register_values(self, product_id: str, *urls: str) -> None:
        if product_id:
            self.product_ids.add(str(product_id))
        self.urls.update(normalized for url in urls if (normalized := normalize_url(url)))

    def save(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        temporary = self.path.with_suffix(".tmp")
        temporary.write_text(
            json.dumps(
                {"urls": sorted(self.urls), "product_ids": sorted(self.product_ids)},
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        temporary.replace(self.path)


def bootstrap_registry_from_excels(directory: Path, registry: ProductRegistry) -> None:
    changed = False
    for path in directory.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                product_id = str(row[1] or "") if len(row) > 1 else ""
                source_url = str(row[6] or "") if len(row) > 6 else ""
                resolved_url = str(row[7] or "") if len(row) > 7 else ""
                before = (len(registry.urls), len(registry.product_ids))
                registry.register_values(product_id, source_url, resolved_url)
                changed = changed or before != (len(registry.urls), len(registry.product_ids))
            workbook.close()
        except (OSError, ValueError, KeyError):
            continue
    if changed:
        registry.save()
